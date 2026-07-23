import io
from collections.abc import Sequence

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.schemas.ticket import TicketOut

CATEGORY_LABELS_UZ = {
    "computer": "Kompyuter",
    "network": "Tarmoq/internet",
    "projector": "Proyektor",
    "power": "Elektr ta'minoti",
    "printer": "Printer",
    "software": "Dasturiy ta'minot",
    "other": "Boshqa",
}
PRIORITY_LABELS_UZ = {"urgent": "Shoshilinch", "normal": "Oddiy"}
STATUS_LABELS_UZ = {"open": "Ochiq", "in_progress": "Jarayonda", "closed": "Yopilgan"}

HEADERS = [
    "Ariza №", "FISH", "Telefon", "Fakultet", "Toifa", "Muhimlik", "Tavsif",
    "Holat", "Texnik xodim", "Yaratilgan sana", "Yopilgan sana", "Yechim izohi",
]
COLUMN_WIDTHS = [14, 22, 16, 22, 16, 12, 40, 12, 22, 18, 18, 40]


def generate_tickets_xlsx(tickets: Sequence[TicketOut]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Arizalar"

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for t in tickets:
        ws.append(
            [
                t.ticket_number,
                t.creator_full_name,
                t.creator_phone,
                t.faculty_name,
                CATEGORY_LABELS_UZ.get(t.category.value, t.category.value),
                PRIORITY_LABELS_UZ.get(t.priority.value, t.priority.value),
                t.description,
                STATUS_LABELS_UZ.get(t.status.value, t.status.value),
                t.technician_full_name or "-",
                t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else "-",
                t.closed_at.strftime("%Y-%m-%d %H:%M") if t.closed_at else "-",
                t.resolution_comment or "-",
            ]
        )

    for index, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
