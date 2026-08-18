import io
from collections.abc import Sequence
from dataclasses import dataclass

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.schemas.inventory import INVENTORY_STATUS_OPTIONS, INVENTORY_TYPE_OPTIONS
from app.schemas.inventory import InventoryItemOut
from app.services.excel_safety import safe_row

HEADERS = [
    "№", "FAKULTET YOKI BO'LIM", "XONA", "INVENTAR RAQAMI",
    "INVENTAR UZASBO", "INVENTAR TOIFASI", "MODELI", "XOLATI",
    "INTERNETGA ULANGANLIGI", "MAC ADRESI", "MA'SUL SHAXS",
]
COLUMN_WIDTHS = [6, 34, 12, 16, 18, 20, 20, 14, 24, 18, 22]

LOCATION_SHEET_NAME = "Lookup"


def faculty_location_label(name: str, parent_name: str | None) -> str:
    return f"{parent_name} / {name}" if parent_name else name


def generate_inventory_xlsx(items: Sequence[InventoryItemOut], location_options: Sequence[str] = ()) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventar"

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for index, item in enumerate(items, start=1):
        ws.append(
            safe_row(
                [
                    index,
                    item.faculty_name,
                    item.room or "",
                    item.inventory_number or "",
                    item.uzasbo or "",
                    item.inventory_type or "",
                    item.model or "",
                    item.status or "",
                    item.internet_connection or "",
                    item.mac_address or "",
                    item.responsible_person or "",
                ]
            )
        )

    for col_index, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_index)].width = width

    max_row = max(len(items) + 1, 500)

    type_validation = DataValidation(
        type="list", formula1=f'"{",".join(INVENTORY_TYPE_OPTIONS)}"', allow_blank=True
    )
    ws.add_data_validation(type_validation)
    type_validation.add(f"F2:F{max_row}")

    status_validation = DataValidation(
        type="list", formula1=f'"{",".join(INVENTORY_STATUS_OPTIONS)}"', allow_blank=True
    )
    ws.add_data_validation(status_validation)
    status_validation.add(f"H2:H{max_row}")

    if location_options:
        lookup_ws = wb.create_sheet(LOCATION_SHEET_NAME)
        for row_index, label in enumerate(location_options, start=1):
            lookup_ws.cell(row=row_index, column=1, value=label)
        lookup_ws.sheet_state = "hidden"

        location_validation = DataValidation(
            type="list",
            formula1=f"={LOCATION_SHEET_NAME}!$A$1:$A${len(location_options)}",
            allow_blank=True,
        )
        ws.add_data_validation(location_validation)
        location_validation.add(f"B2:B{max_row}")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@dataclass
class InventoryImportRow:
    row_number: int
    location_label: str
    room: str | None
    inventory_number: str | None
    uzasbo: str | None
    inventory_type: str | None
    model: str | None
    status: str | None
    internet_connection: str | None
    mac_address: str | None
    responsible_person: str | None


def _clean(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_status(value) -> str:
    cleaned = _clean(value)
    if cleaned is None:
        return "Ishchi"
    return cleaned


def parse_inventory_rows(file_bytes: bytes) -> list[InventoryImportRow]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    rows: list[InventoryImportRow] = []
    for row_number, row in enumerate(ws.iter_rows(min_row=2, max_col=11, values_only=True), start=2):
        padded = list(row) + [None] * (11 - len(row))
        (
            _no,
            location_label,
            room,
            inventory_number,
            uzasbo,
            inventory_type,
            model,
            status,
            internet_connection,
            mac_address,
            responsible_person,
        ) = padded[:11]

        location_label_clean = _clean(location_label)
        if location_label_clean is None:
            continue

        rows.append(
            InventoryImportRow(
                row_number=row_number,
                location_label=location_label_clean,
                room=_clean(room),
                inventory_number=_clean(inventory_number),
                uzasbo=_clean(uzasbo),
                inventory_type=_clean(inventory_type),
                model=_clean(model),
                status=_normalize_status(status),
                internet_connection=_clean(internet_connection),
                mac_address=_clean(mac_address),
                responsible_person=_clean(responsible_person),
            )
        )
    return rows
