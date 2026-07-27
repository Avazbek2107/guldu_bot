import io
from collections.abc import Sequence
from dataclasses import dataclass

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.schemas.faculty import FacultyOut

HEADERS = ["№", "NOMI"]
COLUMN_WIDTHS = [6, 40]


def generate_faculty_xlsx(faculties: Sequence[FacultyOut]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ro'yxat"

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for index, faculty in enumerate(faculties, start=1):
        ws.append([index, faculty.name])

    for col_index, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_index)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@dataclass
class FacultyImportRow:
    row_number: int
    name: str


def _clean(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def parse_faculty_rows(file_bytes: bytes) -> list[FacultyImportRow]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    rows: list[FacultyImportRow] = []
    for row_number, row in enumerate(ws.iter_rows(min_row=2, max_col=2, values_only=True), start=2):
        padded = list(row) + [None] * (2 - len(row))
        _no, name = padded[:2]

        name_clean = _clean(name)
        if name_clean is None:
            continue

        rows.append(FacultyImportRow(row_number=row_number, name=name_clean))
    return rows
